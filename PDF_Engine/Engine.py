import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pdfplumber
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler


# =========================
# PATHS
# =========================
BASE_DIR = Path(__file__).resolve().parent

INPUT_DIR = BASE_DIR / "Input"           # AUTO: GE vagy SAJATKESZLET
INPUT_EK_DIR = BASE_DIR / "Input_EK"     # EK külön input
OUTPUT_DIR = BASE_DIR / "Output"

LOOKUP_FILE = BASE_DIR / "lookup.xlsx"

TEMPLATES_DIR = BASE_DIR / "templates"
EK_TEMPLATE_FILE = TEMPLATES_DIR / "EK_template.xlsx"
GE_TEMPLATE_FILE = TEMPLATES_DIR / "GE_template.xlsx"
SK_TEMPLATE_FILE = TEMPLATES_DIR / "SAJATKESZLET_template.xlsx"

FILL_WARN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


# =========================
# SPEED/ROBUSTNESS
# =========================
def wait_until_file_ready(path: Path, timeout=30, step=0.25):
    """Vár, amíg a fájl másolása befejeződik (méret stabil)."""
    start = time.time()
    last = -1
    stable = 0
    while time.time() - start < timeout:
        try:
            size = path.stat().st_size
        except FileNotFoundError:
            time.sleep(step)
            continue

        if size > 0 and size == last:
            stable += 1
            if stable >= 2:
                return
        else:
            stable = 0
            last = size

        time.sleep(step)


def extract_text_from_pdf_pages(pdf_path: Path, max_pages: int) -> str:
    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            if i >= max_pages:
                break
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def extract_text_from_pdf_all(pdf_path: Path) -> str:
    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


# =========================
# TEMPLATE HELPERS
# =========================
def safe_headers_from_template(ws) -> List[str]:
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers.append(v.strip())
    return headers


def shrink_first_excel_table(ws, last_col_letter: str):
    """Az első Excel Table ref-et a valós sorokig húzza."""
    last_row = ws.max_row
    if last_row < 1:
        return
    new_ref = f"A1:{last_col_letter}{last_row}"

    table_obj = None
    if hasattr(ws, "tables") and ws.tables:
        table_obj = list(ws.tables.values())[0]
    elif hasattr(ws, "_tables") and ws._tables:
        table_obj = list(ws._tables.values())[0]

    if table_obj is not None:
        table_obj.ref = new_ref
        if getattr(table_obj, "autoFilter", None) is not None:
            table_obj.autoFilter.ref = new_ref


# =========================
# COMMON
# =========================
def norm_key(s: str) -> str:
    """GE lookup kulcs: csak betű/szám marad."""
    s = "" if s is None else str(s).strip().upper()
    return re.sub(r"[^0-9A-Z]", "", s)


def parse_float_any(s: str) -> float:
    s = str(s).strip()
    if s.startswith(".") or s.startswith(","):
        s = "0" + s
    return float(s.replace(",", "."))


# ============================================================
# EK + SAJATKESZLET: LOOKUP + ROBUSZTUS POZÍCIÓ-BLOKK PARSER
# ============================================================
EK_LOOKUP_SHEET_NAME = "Sheet1"
EK_LOOKUP_COL_CODE = 2  # B (SAP-Cikkszám)
EK_LOOKUP_COL_NAME = 4  # D (Megnevezés)
EK_LOOKUP_COL_AE = 5    # E (ÁE)
EK_LOOKUP_COL_EK = 8    # H (EK)


def ek_normalize_code(value) -> str:
    """Lookup cikkszám normalizálás (a bevált mintákkal)."""
    if value is None:
        return ""
    tokens = re.findall(r"\d+", str(value))
    if not tokens:
        return ""

    # 6-3-3  (240716 000 200)
    if len(tokens) >= 3 and len(tokens[0]) == 6 and len(tokens[1]) == 3 and len(tokens[2]) == 3:
        return "".join(tokens[:3])

    # >=7-3-4 (040920008 000 1000)
    if len(tokens) >= 3 and len(tokens[0]) >= 7 and len(tokens[1]) == 3 and len(tokens[2]) == 4:
        return "".join(tokens[:3])

    # >=7-3-2/3
    if len(tokens) >= 3 and len(tokens[0]) >= 7 and len(tokens[1]) == 3 and len(tokens[2]) in (2, 3):
        return "".join(tokens[:3])

    # 6-2-3-2
    if len(tokens) >= 4 and len(tokens[0]) == 6 and len(tokens[1]) == 2 and len(tokens[2]) == 3 and len(tokens[3]) == 2:
        return "".join(tokens[:4])

    # 5/6-(1/2)-3-4
    if len(tokens) >= 4 and len(tokens[0]) in (5, 6) and len(tokens[1]) in (1, 2) and len(tokens[2]) == 3 and len(tokens[3]) == 4:
        return "".join(tokens[:4])

    # 5/6-(1/2)-3-3
    if len(tokens) >= 4 and len(tokens[0]) in (5, 6) and len(tokens[1]) in (1, 2) and len(tokens[2]) == 3 and len(tokens[3]) == 3:
        return "".join(tokens[:4])

    # 5-3-4
    if len(tokens) >= 3 and len(tokens[0]) == 5 and len(tokens[1]) == 3 and len(tokens[2]) == 4:
        return "".join(tokens[:3])

    # 5-3
    if len(tokens) >= 2 and len(tokens[0]) == 5 and len(tokens[1]) == 3:
        return "".join(tokens[:2])

    # fallback
    out = []
    for t in tokens:
        if len(t) <= 2 and out:
            break
        out.append(t)
    return "".join(out)


def ek_load_lookup_map() -> Dict[str, dict]:
    """EK/SK lookup cache (Sheet1, oszlopok B/D/E/H)."""
    if not LOOKUP_FILE.exists():
        raise FileNotFoundError(f"Nem találom a lookup fájlt: {LOOKUP_FILE}")

    wb = openpyxl.load_workbook(LOOKUP_FILE, data_only=True)
    ws = wb[EK_LOOKUP_SHEET_NAME] if EK_LOOKUP_SHEET_NAME in wb.sheetnames else wb.active

    start_row = 3
    m: Dict[str, dict] = {}

    for r in range(start_row, ws.max_row + 1):
        sap = ws.cell(r, EK_LOOKUP_COL_CODE).value
        name_hu = ws.cell(r, EK_LOOKUP_COL_NAME).value
        ae = ws.cell(r, EK_LOOKUP_COL_AE).value
        ek = ws.cell(r, EK_LOOKUP_COL_EK).value

        k = ek_normalize_code(sap)
        if not k:
            continue

        m[k] = {
            "code": str(sap).strip() if sap is not None else "",
            "name": str(name_hu).strip() if name_hu is not None else "",
            "AE": ae,
            "EK": ek,
        }

    return m


def build_lookup_lengths(lookup_map: Dict[str, dict]) -> List[int]:
    """Kulcshosszok csökkenő sorrendben (gyors prefix/substring match-hez)."""
    lens = sorted({len(k) for k in lookup_map.keys()}, reverse=True)
    return lens


def best_key_from_digits(digits: str, lookup_map: Dict[str, dict], key_lens: List[int]) -> str:
    """
    Ha a pdfplumber összeragasztja a számokat (pl. 240716000200890...),
    akkor ebből megkeressük a lookupban létező cikkszám kulcsot.
    - először prefix match
    - majd kis eltolással substring (0..6 karakter)
    """
    if not digits:
        return ""

    max_shift = min(6, max(0, len(digits) - 1))
    for shift in range(0, max_shift + 1):
        for L in key_lens:
            if shift + L <= len(digits):
                cand = digits[shift:shift + L]
                if cand in lookup_map:
                    return cand
    return ""


def extract_qty_from_block(block: str) -> Optional[int]:
    """
    Sajátkészlet mennyiség:
    - keressük: 'Darabszám: 123' vagy 'Darabszám 123'
    - plusz kompat: '123 darab' vagy '123 db'
    """
    # 1) Darabszám: 123 / Darabszám 123
    m = re.search(r"Darabszám\s*:?\s*(\d{1,7})\b", block, flags=re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return None

    # 2) régi: 123 darab / 123 db
    m = re.search(r"(\d{1,7})\s*(?:darab|db)\b", block, flags=re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return None

    return None


def ek_parse_items_pos_blocks(pdf_text: str, lookup_map: Dict[str, dict], key_lens: List[int]) -> List[dict]:
    """
    Pozíció-blokk parser + lookup-validált cikkszám kulcs:
    - pozíció kezdet: sor elején 1-3 számjegy
    - blokk = pozíciótól következő pozícióig
    - blokk eleji számokból keresünk olyan kulcsot, ami biztosan létezik a lookupban
    """
    text = pdf_text.replace("\r", "\n")

    pos_re = re.compile(r"(?m)^\s*(\d{1,3})\s+")
    matches = list(pos_re.finditer(text))
    items: List[dict] = []

    if not matches:
        return items

    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        if not block:
            continue

        low = block.lower()
        if low.startswith("poz.") or "vevőszám" in low or low.startswith("megrendelés"):
            continue

        # a blokk elejéből vegyünk egy szeletet (itt szokott lenni a cikkszám)
        head = block[:500]
        digits = "".join(re.findall(r"\d+", head))

        # 1) próbáljuk a normalizálós mintákat a head szövegből
        key_try = ek_normalize_code(head)

        # 2) ha nem jó / nincs a lookupban -> lookup-alapú keresés
        if not key_try or key_try not in lookup_map:
            key_try = best_key_from_digits(digits, lookup_map, key_lens)

        if not key_try:
            continue

        qty = extract_qty_from_block(block)

        items.append({"key": key_try, "qty": qty})

    return items


def ek_extract_customer_no(pdf_text: str) -> str:
    m = re.search(r"Vevőszám:\s*0*(\d+)", pdf_text)
    return m.group(1) if m else "UNKNOWN"


# =========================
# EK OUTPUT
# =========================
def ek_write_output(items: List[dict], out_path: Path, lookup_map: Dict[str, dict]):
    wb = load_workbook(EK_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    headers = safe_headers_from_template(ws)
    if not headers:
        raise ValueError("EK_template.xlsx első sora (header) üres.")

    missing = 0

    for it in items:
        hit = lookup_map.get(it["key"])
        row_data = {}

        if hit:
            row_data["Cikkszám"] = hit["code"]
            row_data["Megnevezés"] = hit["name"]
            row_data["EK"] = hit["EK"]
            row_data["ÁE"] = hit["AE"]
        else:
            missing += 1
            row_data["Cikkszám"] = it["key"]
            row_data["Megnevezés"] = ""
            row_data["EK"] = ""
            row_data["ÁE"] = ""

        ws.append([row_data.get(h, "") for h in headers])

        if not hit:
            row_idx = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row_idx, c).fill = FILL_WARN

        if "Cikkszám" in headers:
            row_idx = ws.max_row
            col_idx = headers.index("Cikkszám") + 1
            ws.cell(row_idx, col_idx).number_format = "@"

    shrink_first_excel_table(ws, last_col_letter=chr(ord("A") + len(headers) - 1))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return missing, len(items)


# =========================
# SAJATKESZLET OUTPUT
# =========================
def sk_write_output(items: List[dict], out_path: Path, lookup_map: Dict[str, dict]):
    wb = load_workbook(SK_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    headers = safe_headers_from_template(ws)
    if not headers:
        raise ValueError("SAJATKESZLET_template.xlsx első sora (header) üres.")

    for it in items:
        hit = lookup_map.get(it["key"])
        row_data = {}

        if hit:
            row_data["Cikkszám"] = hit["code"]
            row_data["Megnevezés"] = hit["name"]
        else:
            row_data["Cikkszám"] = it["key"]
            row_data["Megnevezés"] = ""

        # ✅ mennyiség a pozíció blokkból
        row_data["Mennyiség"] = it.get("qty", "")

        ws.append([row_data.get(h, "") for h in headers])

        if "Cikkszám" in headers:
            row_idx = ws.max_row
            col_idx = headers.index("Cikkszám") + 1
            ws.cell(row_idx, col_idx).number_format = "@"

    shrink_first_excel_table(ws, last_col_letter=chr(ord("A") + len(headers) - 1))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ============================================================
# GE
# ============================================================
GE_LOOKUP_SHEET_NAME = "Sheet2"
GE_LOOKUP_COL_ITEM = "GE cikkszám"
GE_LOOKUP_COL_RECA = "RECA cikkszám"


def ge_load_lookup_map() -> Dict[str, str]:
    wb = load_workbook(LOOKUP_FILE, data_only=False)
    ws = wb[GE_LOOKUP_SHEET_NAME] if GE_LOOKUP_SHEET_NAME in wb.sheetnames else wb.worksheets[1]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c

    if GE_LOOKUP_COL_ITEM not in headers or GE_LOOKUP_COL_RECA not in headers:
        raise ValueError(
            f"GE lookup fejléc hiba. Kell: '{GE_LOOKUP_COL_ITEM}' és '{GE_LOOKUP_COL_RECA}'. Talált: {list(headers.keys())}"
        )

    col_item = headers[GE_LOOKUP_COL_ITEM]
    col_reca = headers[GE_LOOKUP_COL_RECA]

    mapping: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        item_val = ws.cell(r, col_item).value
        reca_val = ws.cell(r, col_reca).value
        if item_val is None:
            continue
        item_str = str(item_val).strip()
        if not item_str:
            continue
        reca_str = "" if reca_val is None else str(reca_val).strip()
        mapping[norm_key(item_str)] = reca_str

    return mapping


def ge_extract_release_number(pdf_text: str) -> str:
    m = re.search(r"\b\d{6,}-\d{1,}\b", pdf_text)
    return m.group(0) if m else "GE_OUTPUT"


def ge_extract_items_from_text(full_text: str) -> List[dict]:
    rows = []
    PRICE_RE = r"(?:\d+(?:[.,]\d+)?|[.,]\d+)"
    re_item = re.compile(r"\bItem:\s*([0-9A-Za-z\-]+)")
    re_each = re.compile(rf"\bEACH\s+(\d+)\s+({PRICE_RE})\s+({PRICE_RE})")

    item_matches = list(re_item.finditer(full_text))
    for i, match in enumerate(item_matches):
        item_code = match.group(1).strip()
        start = match.end()
        end = item_matches[i + 1].start() if i + 1 < len(item_matches) else len(full_text)
        block = full_text[start:end]
        m_each = re_each.search(block)
        if not m_each:
            continue
        qty = int(m_each.group(1))
        price = parse_float_any(m_each.group(2))
        rows.append({"GE cikkszám": item_code, "Mennyiség": qty, "Nettó egységár": price})
    return rows


def ge_write_output(rows: List[dict], out_path: Path, ge_lookup_map: Dict[str, str]):
    wb = load_workbook(GE_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    headers = safe_headers_from_template(ws)
    if not headers:
        raise ValueError("GE_template.xlsx első sora (header) üres.")

    for r in rows:
        ge_code = r.get("GE cikkszám", "")
        reca = ge_lookup_map.get(norm_key(ge_code), "")

        row_data = {
            "GE Cikkszám": ge_code,
            "Reca cikkszám": reca,
            "Mennyiség": r.get("Mennyiség", ""),
            "GE ár": r.get("Nettó egységár", ""),
        }

        ws.append([row_data.get(h, "") for h in headers])

        if "Reca cikkszám" in headers:
            row_idx = ws.max_row
            col_idx = headers.index("Reca cikkszám") + 1
            ws.cell(row_idx, col_idx).number_format = "@"

    shrink_first_excel_table(ws, last_col_letter=chr(ord("A") + len(headers) - 1))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ============================================================
# TYPE DETECT (quick)
# ============================================================
def is_ge_pdf_quick(first_pages_text: str) -> bool:
    low = first_pages_text.lower()
    score = 0
    if "item:" in low:
        score += 1
    if "each" in low:
        score += 1
    if "net unit price" in low:
        score += 1
    if "blanket release" in low:
        score += 1
    return score >= 2


# ============================================================
# PROCESSORS
# ============================================================
def process_pdf_ek(pdf_path: Path, lookup_map: Dict[str, dict], key_lens: List[int]):
    full_text = extract_text_from_pdf_all(pdf_path)
    customer_no = ek_extract_customer_no(full_text)

    items = ek_parse_items_pos_blocks(full_text, lookup_map, key_lens)
    if not items:
        print(f"[EK][WARN] 0 tétel: {pdf_path.name}")
        return

    out_path = OUTPUT_DIR / f"{customer_no} EK.xlsx"
    missing, total = ek_write_output(items, out_path, lookup_map)
    print(f"[EK][OK] {pdf_path.name} -> {out_path.name} | tételek: {total} | nem talált: {missing}")


def process_pdf_sk(pdf_path: Path, full_text: str, lookup_map: Dict[str, dict], key_lens: List[int]):
    items = ek_parse_items_pos_blocks(full_text, lookup_map, key_lens)
    if not items:
        print(f"[SK][WARN] 0 tétel: {pdf_path.name}")
        return

    out_path = OUTPUT_DIR / f"SAJATKESZLET_{pdf_path.stem}.xlsx"
    sk_write_output(items, out_path, lookup_map)
    print(f"[SK][OK] {pdf_path.name} -> {out_path.name} | sorok: {len(items)}")


def process_pdf_ge(pdf_path: Path, full_text: str, ge_lookup_map: Dict[str, str]):
    rows = ge_extract_items_from_text(full_text)
    if not rows:
        print(f"[GE][WARN] 0 tétel: {pdf_path.name}")
        return

    release_no = ge_extract_release_number(full_text)
    out_path = OUTPUT_DIR / f"{release_no}.xlsx"
    ge_write_output(rows, out_path, ge_lookup_map)
    print(f"[GE][OK] {pdf_path.name} -> {out_path.name} | sorok: {len(rows)}")


# ============================================================
# WATCHDOG
# ============================================================
class EkHandler(FileSystemEventHandler):
    def __init__(self, lookup_map: Dict[str, dict], key_lens: List[int]):
        self.lookup_map = lookup_map
        self.key_lens = key_lens

    def on_created(self, event):
        if event.is_directory:
            return
        p = Path(event.src_path)
        if p.suffix.lower() != ".pdf":
            return

        wait_until_file_ready(p)
        try:
            process_pdf_ek(p, self.lookup_map, self.key_lens)
        except Exception as e:
            print(f"[EK][ERROR] {p.name}: {e}")


class AutoHandler(FileSystemEventHandler):
    def __init__(self, lookup_map: Dict[str, dict], key_lens: List[int], ge_lookup_map: Dict[str, str]):
        self.lookup_map = lookup_map
        self.key_lens = key_lens
        self.ge_lookup_map = ge_lookup_map

    def on_created(self, event):
        if event.is_directory:
            return
        p = Path(event.src_path)
        if p.suffix.lower() != ".pdf":
            return

        wait_until_file_ready(p)

        try:
            head_text = extract_text_from_pdf_pages(p, max_pages=2)
            full_text = extract_text_from_pdf_all(p)

            if is_ge_pdf_quick(head_text):
                process_pdf_ge(p, full_text, self.ge_lookup_map)
            else:
                process_pdf_sk(p, full_text, self.lookup_map, self.key_lens)

        except Exception as e:
            print(f"[AUTO][ERROR] {p.name}: {e}")


# ============================================================
# MAIN
# ============================================================
def main():
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_EK_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for t in [EK_TEMPLATE_FILE, GE_TEMPLATE_FILE, SK_TEMPLATE_FILE]:
        if not t.exists():
            print(f"[HIBA] Hiányzik template: {t}")
            return

    if not LOOKUP_FILE.exists():
        print(f"[HIBA] Hiányzik lookup: {LOOKUP_FILE}")
        return

    print("[INFO] Lookup cache betöltése...")
    ek_lookup_map = ek_load_lookup_map()
    key_lens = build_lookup_lengths(ek_lookup_map)

    ge_lookup_map = ge_load_lookup_map()

    print(f"[INFO] EK/SK lookup: {len(ek_lookup_map)} | GE lookup: {len(ge_lookup_map)}")

    observer = Observer()
    observer.schedule(EkHandler(ek_lookup_map, key_lens), str(INPUT_EK_DIR), recursive=False)
    observer.schedule(AutoHandler(ek_lookup_map, key_lens, ge_lookup_map), str(INPUT_DIR), recursive=False)
    observer.start()

    print(f"Watching EK:   {INPUT_EK_DIR}")
    print(f"Watching AUTO: {INPUT_DIR} (GE vs SAJATKESZLET)")
    print(f"Output:        {OUTPUT_DIR}")
    print("Kilépés: Ctrl+C")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()